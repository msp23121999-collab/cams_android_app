import os
import json
import csv
import shutil
import zipfile
import logging
from datetime import datetime, date, timezone, time
from decimal import Decimal
from typing import List, Dict, Any, Optional
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text, delete, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from xhtml2pdf import pisa

from app.db.base import Base
from app.db.models.backup import BackupConfiguration, BackupHistory
from app.db.models.audit import AuditLog
from app.db.models.user import User, UserRole
logger = logging.getLogger("app.backup_service")

from app.db.models import (
    Student, Department, Degree, Attendance, Section, Course,
    Payment, FeeRecord, FeeStructure, Grievance, FacultyProfile,
    StudyMaterial, Assignment
)

BACKUP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "backups"))
UPLOAD_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "uploads"))

class BackupService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_config(self) -> BackupConfiguration:
        result = await self.db.execute(select(BackupConfiguration).where(BackupConfiguration.is_deleted.is_(False)))
        config = result.scalar_one_or_none()
        if not config:
            config = BackupConfiguration(
                auto_backup_enabled=True,
                schedule_time="21:00",
                retention_count=30
            )
            self.db.add(config)
            await self.db.commit()
            await self.db.refresh(config)
        return config

    async def update_config(self, enabled: bool, schedule_time: str, retention_count: int) -> BackupConfiguration:
        config = await self.get_config()
        config.auto_backup_enabled = enabled
        config.schedule_time = schedule_time
        config.retention_count = retention_count
        await self.db.commit()
        await self.db.refresh(config)
        return config

    async def log_audit(self, action: str, entity_id: Optional[str], user_id: Optional[str] = None):
        audit = AuditLog(
            user_id=user_id,
            action=action,
            entity="BackupRestore",
            entity_id=entity_id,
            ip_address="127.0.0.1",
            timestamp=datetime.now(timezone.utc)
        )
        self.db.add(audit)
        await self.db.commit()

    async def create_backup(self, trigger_type: str = "MANUAL", is_incremental: bool = False, user_id: Optional[str] = None) -> BackupHistory:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
        filename = f"backup_{timestamp}.zip"
        filepath = os.path.join(BACKUP_DIR, filename)

        # Get last successful backup time for incremental diffs
        last_backup_time = None
        if is_incremental:
            last_successful = await self.db.execute(
                select(BackupHistory)
                .where(BackupHistory.status == "SUCCESS")
                .order_by(BackupHistory.created_at.desc())
                .limit(1)
            )
            last_hist = last_successful.scalar_one_or_none()
            if last_hist:
                last_backup_time = last_hist.created_at
            else:
                # No previous backup exists, fall back to full backup
                is_incremental = False

        temp_dir = os.path.join(BACKUP_DIR, f"temp_{timestamp}")
        os.makedirs(temp_dir, exist_ok=True)

        try:
            # Metadata description
            metadata = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "is_incremental": is_incremental,
                "last_backup_time": last_backup_time.isoformat() if last_backup_time else None,
                "trigger_type": trigger_type,
                "tables": []
            }

            # Export database tables
            db_dir = os.path.join(temp_dir, "db")
            csv_dir = os.path.join(temp_dir, "reports_csv")
            pdf_dir = os.path.join(temp_dir, "reports_pdf")
            os.makedirs(db_dir, exist_ok=True)
            os.makedirs(csv_dir, exist_ok=True)
            os.makedirs(pdf_dir, exist_ok=True)

            # Build FK mappings dynamically for user-friendly exports
            fk_mappings = {}
            try:
                users_res = await self.db.execute(select(User.id, User.full_name))
                fk_mappings["users"] = {r[0]: r[1] for r in users_res.all()}

                students_res = await self.db.execute(
                    select(Student.id, Student.roll_no, User.full_name)
                    .join(User, Student.user_id == User.id)
                )
                fk_mappings["students"] = {r[0]: f"{r[2]} ({r[1]})" for r in students_res.all()}

                faculty_res = await self.db.execute(
                    select(FacultyProfile.id, User.full_name)
                    .join(User, FacultyProfile.user_id == User.id)
                )
                fk_mappings["faculty_profiles"] = {r[0]: r[1] for r in faculty_res.all()}

                # Auto reflect and build descriptive mappings for other tables
                for t in Base.metadata.sorted_tables:
                    if t.name in ["users", "students", "faculty_profiles", "backup_configurations", "backup_history", "audit_logs", "activity_logs"]:
                        continue
                    label_col = None
                    for cname in ["name", "title", "section_name", "code"]:
                        if cname in t.columns:
                            label_col = t.columns[cname]
                            break
                    if label_col is not None:
                        pk_col = list(t.primary_key.columns)[0]
                        rows_res = await self.db.execute(select(pk_col, label_col))
                        fk_mappings[t.name] = {row[0]: row[1] for row in rows_res.all()}
            except Exception as e:
                logger.error(f"Error building FK mappings: {e}")

            # Retrieve tables in topological order
            sorted_tables = Base.metadata.sorted_tables

            for table in sorted_tables:
                table_name = table.name
                query = select(table)
                
                # Check for timestamp mixin fields to filter for incremental backup
                has_timestamp_fields = "created_at" in table.columns and "updated_at" in table.columns
                
                if is_incremental and last_backup_time and has_timestamp_fields:
                    query = query.where(
                        (table.c.updated_at > last_backup_time) | 
                        (table.c.created_at > last_backup_time)
                    )

                result = await self.db.execute(query)
                rows = result.fetchall()

                # Save raw json for restore
                table_data = []
                headers = [col.name for col in table.columns]
                csv_rows = []

                for r in rows:
                    row_dict = {}
                    csv_row = {}
                    for col in table.columns:
                        val = getattr(r, col.name)
                        # JSON serialization formatting (raw values for DB restore)
                        if isinstance(val, (datetime, date, time)):
                            row_dict[col.name] = val.isoformat()
                        elif isinstance(val, Decimal):
                            row_dict[col.name] = float(val)
                        elif isinstance(val, bytes):
                            row_dict[col.name] = val.hex()
                        else:
                            row_dict[col.name] = val

                        # CSV formatting with friendly headers and descriptive foreign keys
                        display_val = val
                        if col.name == "hashed_password":
                            display_val = "Password@123"
                        elif col.foreign_keys and val:
                            fk = list(col.foreign_keys)[0]
                            tgt_table = fk.column.table.name
                            if tgt_table in fk_mappings and val in fk_mappings[tgt_table]:
                                display_val = fk_mappings[tgt_table][val]

                        friendly_h = col.name.replace("_id", " Name").replace("_", " ").title()
                        if friendly_h == "Roll No":
                            friendly_h = "Register Number"
                        
                        csv_row[friendly_h] = str(display_val) if display_val is not None else ""

                    table_data.append(row_dict)
                    csv_rows.append(csv_row)

                # Generate friendly headers list for CSV DictWriter
                friendly_headers = []
                for col in table.columns:
                    fh = col.name.replace("_id", " Name").replace("_", " ").title()
                    if fh == "Roll No":
                        fh = "Register Number"
                    friendly_headers.append(fh)

                # Only save json files if there are rows (or save empty list for full backups)
                if table_data or not is_incremental:
                    json_path = os.path.join(db_dir, f"{table_name}.json")
                    with open(json_path, "w", encoding="utf-8") as f:
                        json.dump(table_data, f, indent=2)

                    # Export user-readable CSV detail sheets
                    csv_path = os.path.join(csv_dir, f"{table_name}_details.csv")
                    with open(csv_path, "w", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(f, fieldnames=friendly_headers)
                        writer.writeheader()
                        writer.writerows(csv_rows)

                    metadata["tables"].append({
                        "name": table_name,
                        "row_count": len(table_data)
                    })

                    # If this is fee or student data, let's also generate a beautiful PDF report
                    if table_name in ["fee_records", "students", "faculty_profiles", "users"] and len(table_data) > 0:
                        pdf_path = os.path.join(pdf_dir, f"{table_name}_report.pdf")
                        await self.generate_pdf_report(table_name, table_data, headers, pdf_path, fk_mappings)

            # Generate the user-friendly Excel workbook
            excel_path = os.path.join(temp_dir, "cams_system_backup.xlsx")
            await self.generate_excel_backup(excel_path)

            # Save metadata.json
            with open(os.path.join(temp_dir, "metadata.json"), "w", encoding="utf-8") as f:
                json.dump(metadata, f, indent=2)

            # Copy uploads folder
            uploads_dest = os.path.join(temp_dir, "uploads")
            if os.path.exists(UPLOAD_DIR):
                shutil.copytree(UPLOAD_DIR, uploads_dest, dirs_exist_ok=True)
            else:
                os.makedirs(uploads_dest, exist_ok=True)

            # Package into ZIP file
            with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zipf:
                for root, _, files in os.walk(temp_dir):
                    for file in files:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(full_path, temp_dir)
                        zipf.write(full_path, rel_path)

            size_bytes = os.path.getsize(filepath)

            # Create backup history record
            history = BackupHistory(
                filename=filename,
                filepath=filepath,
                size_bytes=size_bytes,
                status="SUCCESS",
                trigger_type=trigger_type,
                created_by=user_id,
                is_incremental=is_incremental
            )
            self.db.add(history)
            await self.db.commit()
            await self.db.refresh(history)

            # Log audit
            await self.log_audit(f"Backup Created ({trigger_type})", history.id, user_id)

            # Perform retention rotation
            await self.rotate_backups()

            return history

        except Exception as e:
            # Create failed backup record
            history = BackupHistory(
                filename=filename,
                filepath=filepath,
                size_bytes=0,
                status="FAILED",
                trigger_type=trigger_type,
                created_by=user_id,
                error_message=str(e),
                is_incremental=is_incremental
            )
            self.db.add(history)
            await self.db.commit()
            await self.db.refresh(history)
            
            await self.log_audit(f"Backup Creation Failed: {str(e)[:100]}", history.id, user_id)
            raise e
        finally:
            # Cleanup temp directory
            shutil.rmtree(temp_dir, ignore_errors=True)

    async def generate_pdf_report(self, table_name: str, data: List[Dict], headers: List[str], dest_path: str, fk_mappings: dict = None):
        title = table_name.replace("_", " ").title()
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Helvetica, Arial, sans-serif; font-size: 10px; color: #333; }}
                h1 {{ color: #4F46E5; font-size: 18px; border-bottom: 2px solid #E5E7EB; padding-bottom: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                th {{ background-color: #F3F4F6; color: #111827; border: 1px solid #D1D5DB; padding: 6px; font-weight: bold; }}
                td {{ border: 1px solid #E5E7EB; padding: 6px; }}
                .footer {{ text-align: right; margin-top: 20px; font-size: 8px; color: #9CA3AF; }}
            </style>
        </head>
        <body>
            <h1>CAMS System Backup - {title}</h1>
            <p>Generated at: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
            <table>
                <thead>
                    <tr>
        """
        for h in headers[:7]:  # limit columns to prevent overflow
            friendly_h = h.replace("_id", " Name").replace("_", " ").title()
            if friendly_h == "Roll No":
                friendly_h = "Register Number"
            html += f"<th>{friendly_h.upper()}</th>"
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        for row in data[:100]:  # limit rows for report summary
            html += "<tr>"
            for h in headers[:7]:
                val = row.get(h, "")
                if h == "hashed_password" and val:
                    val = "Password@123"
                elif h.endswith("_id") and val and fk_mappings:
                    found_name = None
                    for tname, tmap in fk_mappings.items():
                        if val in tmap:
                            found_name = tmap[val]
                            break
                    if found_name:
                        val = found_name
                html += f"<td>{val}</td>"
            html += "</tr>"
            
        html += f"""
                </tbody>
            </table>
            <p class="footer">Total Records in Export: {len(data)} (Summary shows up to 100)</p>
        </body>
        </html>
        """
        try:
            with open(dest_path, "wb") as f:
                pisa.CreatePDF(html, dest=f)
        except Exception as e:
            logger.error(f"PDF generation failed for table {table_name}: {e}", exc_info=True)

    async def generate_excel_backup(self, dest_path: str):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # 1. Fetch Students
        students_res = await self.db.execute(
            select(Student, User, Department, Degree)
            .join(User, Student.user_id == User.id)
            .join(Department, Student.department_id == Department.id)
            .outerjoin(Degree, Student.degree_id == Degree.id)
            .where(Student.is_deleted.is_(False))
        )
        students_rows = students_res.all()

        # 2. Fetch Staff
        staff_res = await self.db.execute(
            select(User, FacultyProfile, Department)
            .join(FacultyProfile, FacultyProfile.user_id == User.id)
            .outerjoin(Department, User.department_id == Department.id)
            .where(User.role == UserRole.FACULTY, User.is_deleted.is_(False))
        )
        staff_rows = staff_res.all()

        # 3. Fetch Attendance
        attendance_res = await self.db.execute(
            select(Attendance, Student, User, Section, Course, Department)
            .join(Student, Attendance.student_id == Student.id)
            .join(User, Student.user_id == User.id)
            .join(Section, Attendance.section_id == Section.id)
            .join(Course, Section.course_id == Course.id)
            .join(Department, Student.department_id == Department.id)
            .where(Attendance.is_deleted.is_(False))
        )
        attendance_rows = attendance_res.all()

        # 4. Fetch Fees
        fee_res = await self.db.execute(
            select(FeeRecord, Student, User, FeeStructure, Department)
            .join(Student, FeeRecord.student_id == Student.id)
            .join(User, Student.user_id == User.id)
            .join(FeeStructure, FeeRecord.fee_structure_id == FeeStructure.id)
            .join(Department, Student.department_id == Department.id)
            .where(FeeRecord.is_deleted.is_(False))
        )
        fee_rows = fee_res.all()

        from collections import defaultdict
        student_fees = defaultdict(lambda: {
            "name": "", "roll_no": "", "dept": "", "batch": "",
            "total_fee": 0.0, "paid_amount": 0.0, "last_payment_date": None
        })

        for fr, s, u, fs, dept in fee_rows:
            sf = student_fees[s.id]
            sf["name"] = u.full_name
            sf["roll_no"] = s.roll_no
            sf["dept"] = dept.name
            sf["batch"] = str(s.batch_year)
            sf["total_fee"] += float(fs.amount)

            pay_res = await self.db.execute(
                select(Payment).where(Payment.fee_record_id == fr.id, Payment.is_deleted.is_(False))
            )
            payments = pay_res.scalars().all()
            for p in payments:
                sf["paid_amount"] += float(p.amount)
                if p.paid_at:
                    if not sf["last_payment_date"] or p.paid_at > sf["last_payment_date"]:
                        sf["last_payment_date"] = p.paid_at

        fees_data = []
        for sid, sf in student_fees.items():
            due = max(0.0, sf["total_fee"] - sf["paid_amount"])
            fees_data.append({
                "Student Name": sf["name"],
                "Register Number": sf["roll_no"],
                "Course": sf["dept"],
                "Batch": sf["batch"],
                "Total Fee": sf["total_fee"],
                "Paid Amount": sf["paid_amount"],
                "Due Amount": due,
                "Last Payment Date": sf["last_payment_date"].strftime("%Y-%m-%d %H:%M:%S") if sf["last_payment_date"] else "N/A"
            })

        # 5. Fetch Grievances
        grievance_res = await self.db.execute(
            select(Grievance, User)
            .join(User, Grievance.raised_by == User.id)
            .where(Grievance.is_deleted.is_(False))
        )
        grievance_rows = grievance_res.all()

        # 6. Fetch Documents
        docs_data = []
        sm_res = await self.db.execute(
            select(StudyMaterial, User, Section)
            .join(User, StudyMaterial.faculty_id == User.id)
            .join(Section, StudyMaterial.section_id == Section.id)
            .where(StudyMaterial.is_deleted.is_(False))
        )
        for sm, u, sec in sm_res.all():
            docs_data.append({
                "Document ID": sm.id,
                "Title": sm.title,
                "Type": sm.type or "Lecture Slides",
                "Course/Section": sec.section_name,
                "Uploaded By": u.full_name,
                "File URL/Path": sm.file_url,
                "Uploaded At": sm.created_at.strftime("%Y-%m-%d %H:%M:%S") if sm.created_at else "N/A"
            })

        assign_res = await self.db.execute(
            select(Assignment, User, Section)
            .join(User, Assignment.faculty_id == User.id)
            .join(Section, Assignment.section_id == Section.id)
            .where(Assignment.is_deleted.is_(False))
        )
        for assign, u, sec in assign_res.all():
            docs_data.append({
                "Document ID": assign.id,
                "Title": assign.title,
                "Type": "Assignment",
                "Course/Section": sec.section_name,
                "Uploaded By": u.full_name,
                "File URL/Path": f"Deadline: {assign.deadline.isoformat() if assign.deadline else 'N/A'}",
                "Uploaded At": assign.created_at.strftime("%Y-%m-%d %H:%M:%S") if assign.created_at else "N/A"
            })

        # Calculate counts
        total_courses_q = await self.db.execute(select(sa.func.count(Course.id)).where(Course.is_deleted.is_(False)))
        total_courses = total_courses_q.scalar_one_or_none() or 0

        total_depts_q = await self.db.execute(select(sa.func.count(Department.id)).where(Department.is_deleted.is_(False)))
        total_departments = total_depts_q.scalar_one_or_none() or 0

        # Construct sheets
        data_sheets = {}

        # Students and Admissions
        students_sheet = []
        admissions_sheet = []
        for s, u, dept, reg in students_rows:
            students_sheet.append({
                "Register Number": s.roll_no,
                "Student Name": u.full_name,
                "Gender": "N/A",
                "Course": reg.name if reg else "N/A",
                "Department": dept.name,
                "Batch": reg.applicable_batch if reg else str(s.batch_year),
                "Semester": s.semester,
                "Section": "A",
                "Mobile Number": u.phone or "N/A",
                "Email": u.email
            })
            admissions_sheet.append({
                "Register Number": s.roll_no,
                "Student Name": u.full_name,
                "Department": dept.name,
                "Batch": reg.applicable_batch if reg else str(s.batch_year),
                "Quota": s.quota or "Government",
                "Email": u.email,
                "Phone": u.phone or "N/A",
                "Admission Date": s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "N/A"
            })
        data_sheets["Students"] = students_sheet
        data_sheets["Admissions"] = admissions_sheet

        # Staff Sheet
        staff_sheet = []
        for u, fp, dept in staff_rows:
            staff_sheet.append({
                "Staff ID": u.id,
                "Staff Name": u.full_name,
                "Designation": fp.designation if fp else "Faculty",
                "Department": dept.name if dept else "N/A",
                "Mobile Number": u.phone or "N/A",
                "Email": u.email
            })
        data_sheets["Staff"] = staff_sheet

        # Attendance Sheet
        attendance_sheet = []
        for att, s, u, sec, course, dept in attendance_rows:
            attendance_sheet.append({
                "Student Register Number": s.roll_no,
                "Student Name": u.full_name,
                "Course": course.name,
                "Department": dept.name,
                "Batch": str(s.batch_year),
                "Semester": s.semester,
                "Section": sec.section_name,
                "Attendance Date": att.date.isoformat() if att.date else "N/A",
                "Attendance Status": att.status.value if hasattr(att.status, 'value') else str(att.status)
            })
        data_sheets["Attendance"] = attendance_sheet
        data_sheets["Fees"] = fees_data

        # Grievances Sheet
        grievances_sheet = []
        for g, u in grievance_rows:
            assigned_name = "Unassigned"
            if g.assigned_to:
                assignee_res = await self.db.execute(select(User).where(User.id == g.assigned_to))
                assignee = assignee_res.scalar_one_or_none()
                if assignee:
                    assigned_name = assignee.full_name
            grievances_sheet.append({
                "Grievance ID": g.id,
                "Raised By": u.full_name,
                "Category": g.category,
                "Description": g.description,
                "Status": g.status,
                "Assigned To": assigned_name,
                "Created At": g.created_at.strftime("%Y-%m-%d %H:%M:%S") if g.created_at else "N/A"
            })
        data_sheets["Grievances"] = grievances_sheet
        data_sheets["Documents"] = docs_data

        # Build Excel Workbook
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1F2937")

        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # Write Title
        ws_summary["A1"] = "CAMS System Backup Summary"
        ws_summary["A1"].font = title_font

        ws_summary["A3"] = "Metric"
        ws_summary["B3"] = "Value"
        ws_summary["A3"].fill = header_fill
        ws_summary["A3"].font = header_font
        ws_summary["B3"].fill = header_fill
        ws_summary["B3"].font = header_font

        metrics = [
            ("Total Students", len(students_sheet)),
            ("Total Staff", len(staff_sheet)),
            ("Total Courses", total_courses),
            ("Total Departments", total_departments),
            ("Total Attendance Records", len(attendance_sheet)),
            ("Total Fee Records", len(fee_rows)),
            ("Backup Date & Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for idx, (metric, val) in enumerate(metrics, start=4):
            ws_summary.cell(row=idx, column=1, value=metric).font = Font(name="Segoe UI", size=10, bold=True)
            ws_summary.cell(row=idx, column=2, value=val).font = cell_font
            ws_summary.cell(row=idx, column=1).border = thin_border
            ws_summary.cell(row=idx, column=2).border = thin_border

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 25

        for sheet_name, rows in data_sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            if not rows:
                ws.cell(row=1, column=1, value="No records found").font = Font(name="Segoe UI", size=10, italic=True)
                continue

            headers = list(rows[0].keys())
            # Header Row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data Rows
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    val = row_data[header]
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = cell_font
                    cell.border = thin_border

            # Autofit column widths
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(dest_path)

    async def rotate_backups(self):
        config = await self.get_config()
        retention = config.retention_count

        result = await self.db.execute(
            select(BackupHistory)
            .where(BackupHistory.status == "SUCCESS")
            .order_by(BackupHistory.created_at.desc())
        )
        backups = result.scalars().all()

        if len(backups) > retention:
            to_delete = backups[retention:]
            for b in to_delete:
                # Remove file from disk
                if os.path.exists(b.filepath):
                    try:
                        os.remove(b.filepath)
                    except OSError:
                        pass
                
                # Mark as deleted in db
                b.is_deleted = True
                b.deleted_at = datetime.now(timezone.utc)
                await self.db.commit()
                await self.log_audit("Backup Deleted (Retention)", b.id)

    async def restore_backup(self, backup_id: str, user_id: Optional[str] = None):
        # Fetch history
        result = await self.db.execute(select(BackupHistory).where(BackupHistory.id == backup_id))
        history = result.scalar_one_or_none()
        if not history:
            raise ValueError("Backup not found")
        if not os.path.exists(history.filepath):
            raise FileNotFoundError("Backup archive file not found on disk")

        await self.restore_from_zip_file(history.filepath, user_id)

    async def restore_from_zip_file(self, filepath: str, user_id: Optional[str] = None):
        # Extract to a temp location
        timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        temp_dir = os.path.join(BACKUP_DIR, f"restore_temp_{timestamp}")
        os.makedirs(temp_dir, exist_ok=True)

        try:
            with zipfile.ZipFile(filepath, "r") as zipf:
                zipf.extractall(temp_dir)

            # Read metadata
            metadata_path = os.path.join(temp_dir, "metadata.json")
            if not os.path.exists(metadata_path):
                raise ValueError("Invalid backup file: metadata.json missing")

            with open(metadata_path, "r", encoding="utf-8") as f:
                metadata = json.load(f)

            is_incremental = metadata.get("is_incremental", False)

            # If it's incremental, we must find the base full backup and intermediate incrementals
            # to reconstruct the full state.
            chain = []
            if is_incremental:
                # Retrieve all successful backups up to the current timestamp
                backups_res = await self.db.execute(
                    select(BackupHistory)
                    .where(BackupHistory.status == "SUCCESS")
                    .order_by(BackupHistory.created_at.asc())
                )
                all_backups = backups_res.scalars().all()

                # Find our backup file in history
                tgt_idx = -1
                for idx, b in enumerate(all_backups):
                    if b.filepath == filepath or os.path.basename(b.filepath) == os.path.basename(filepath):
                        tgt_idx = idx
                        break

                if tgt_idx == -1:
                    raise ValueError("Incremental backup not found in system history. Cannot rebuild chain.")

                # Go backwards to find the nearest full backup
                full_idx = -1
                for idx in range(tgt_idx, -1, -1):
                    if not all_backups[idx].is_incremental:
                        full_idx = idx
                        break

                if full_idx == -1:
                    raise ValueError("Could not find a base full backup for this incremental backup.")

                # The chain contains the full backup and all incrementals up to tgt_idx
                chain = all_backups[full_idx:tgt_idx + 1]
            else:
                # Full backup: just restore this one
                chain = [filepath]

            # Reconstruct the database rows
            # Topological order
            sorted_tables = Base.metadata.sorted_tables

            # Disable constraints check for postgres
            await self.db.execute(text("SET session_replication_role = 'replica';"))

            # Clean database first
            for table in reversed(sorted_tables):
                # Don't truncate configuration/history/audit tables so we don't lock ourselves out of settings/audit
                if table.name in ["backup_configurations", "backup_history", "audit_logs", "activity_logs"]:
                    continue
                await self.db.execute(delete(table))
            await self.db.commit()

            # Now, apply each backup in the chain
            for item in chain:
                # If item is a string, it's the direct zip filepath. If it's a BackupHistory, it's a BackupHistory object.
                item_zip_path = item if isinstance(item, str) else item.filepath
                
                # Unzip item to a sub-temp directory
                sub_temp = os.path.join(temp_dir, f"sub_{os.path.basename(item_zip_path)}")
                os.makedirs(sub_temp, exist_ok=True)
                with zipfile.ZipFile(item_zip_path, "r") as zipf:
                    zipf.extractall(sub_temp)

                # Process tables
                for table in sorted_tables:
                    table_name = table.name
                    if table_name in ["backup_configurations", "backup_history", "audit_logs", "activity_logs"]:
                        continue

                    json_path = os.path.join(sub_temp, "db", f"{table_name}.json")
                    if not os.path.exists(json_path):
                        continue

                    with open(json_path, "r", encoding="utf-8") as f:
                        records = json.load(f)

                    if not records:
                        continue

                    # Bulk deserialize records
                    deserialized_records = []
                    for rec in records:
                        deserialized_rec = {}
                        for col in table.columns:
                            val = rec.get(col.name, None)
                            # Parse ISO datetime/date
                            if val is not None:
                                if isinstance(col.type, sa.DateTime):
                                    deserialized_rec[col.name] = datetime.fromisoformat(val)
                                elif isinstance(col.type, sa.Date):
                                    deserialized_rec[col.name] = date.fromisoformat(val)
                                elif isinstance(col.type, sa.Time):
                                    deserialized_rec[col.name] = time.fromisoformat(val)
                                elif isinstance(col.type, sa.Boolean):
                                    deserialized_rec[col.name] = bool(val)
                                else:
                                    deserialized_rec[col.name] = val
                            else:
                                # Fallback default values for NOT NULL columns
                                if not col.nullable and not col.primary_key:
                                    if col.name == "quota":
                                        deserialized_rec[col.name] = "Government"
                                    elif isinstance(col.type, (sa.Integer, sa.Numeric, sa.Float)):
                                        deserialized_rec[col.name] = 0
                                    elif isinstance(col.type, sa.Boolean):
                                        deserialized_rec[col.name] = False
                                    elif isinstance(col.type, sa.DateTime):
                                        deserialized_rec[col.name] = datetime.now()
                                    elif isinstance(col.type, sa.Date):
                                        deserialized_rec[col.name] = date.today()
                                    elif isinstance(col.type, sa.Time):
                                        deserialized_rec[col.name] = time(0, 0, 0)
                                    elif isinstance(col.type, sa.String):
                                        deserialized_rec[col.name] = ""
                                    else:
                                        deserialized_rec[col.name] = ""
                                else:
                                    deserialized_rec[col.name] = None
                        deserialized_records.append(deserialized_rec)

                    if deserialized_records:
                        pk_col = table.primary_key.columns.values()[0]
                        stmt = pg_insert(table).values(deserialized_records)
                        
                        # We update all columns except the primary key columns
                        update_cols = {col.name: stmt.excluded[col.name] for col in table.columns if not col.primary_key}
                        
                        if update_cols:
                            upsert_stmt = stmt.on_conflict_do_update(
                                index_elements=[pk_col],
                                set_=update_cols
                            )
                        else:
                            upsert_stmt = stmt.on_conflict_do_nothing(index_elements=[pk_col])
                            
                        await self.db.execute(upsert_stmt)

                # Copy uploads folder from the final item in the chain or overlay them
                uploads_src = os.path.join(sub_temp, "uploads")
                if os.path.exists(uploads_src):
                    os.makedirs(UPLOAD_DIR, exist_ok=True)
                    shutil.copytree(uploads_src, UPLOAD_DIR, dirs_exist_ok=True)

                shutil.rmtree(sub_temp, ignore_errors=True)

            # Reset constraints check for postgres
            await self.db.execute(text("SET session_replication_role = 'origin';"))
            await self.db.commit()

            # Log audit
            await self.log_audit("Backup Restored Successfully", os.path.basename(filepath), user_id)

        except Exception as e:
            try:
                await self.db.execute(text("SET session_replication_role = 'origin';"))
            except Exception:
                pass
            await self.db.rollback()
            await self.log_audit(f"Restore Failed: {str(e)[:100]}", os.path.basename(filepath), user_id)
            raise e
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
